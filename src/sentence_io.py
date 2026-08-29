"""导入已有例句（与导出格式互认）：Excel / JSON / Markdown。

导入后可直接进入复习模式，无需重新调用大模型生成。
"""

from __future__ import annotations

import json
import re
from io import BytesIO
from pathlib import Path

from .generator import SentenceItem


def _make_item(
    word: object,
    english: object,
    chinese: object = "",
    sense: object = "",
    index: object = 1,
) -> SentenceItem | None:
    """校验并构造 SentenceItem；word 和 english 缺一不可。"""
    w = str(word or "").strip()
    en = str(english or "").strip()
    if not w or not en:
        return None
    try:
        idx = max(1, int(index))
    except (TypeError, ValueError):
        idx = 1
    return SentenceItem(
        word=w,
        english=en,
        chinese=str(chinese or "").strip(),
        sense=str(sense or "").strip(),
        index=idx,
    )


def import_sentences_json(content: bytes) -> list[SentenceItem]:
    """解析导出的 JSON：[{word, english, chinese, sense, index}, ...]。"""
    try:
        data = json.loads(content.decode("utf-8-sig"))
    except (UnicodeDecodeError, json.JSONDecodeError) as e:
        raise ValueError(f"JSON 解析失败：{e}") from e
    if isinstance(data, dict):
        data = data.get("sentences") or data.get("items") or []
    if not isinstance(data, list):
        raise ValueError("JSON 结构不正确：应为例句对象数组")

    items: list[SentenceItem] = []
    for row in data:
        if isinstance(row, dict):
            it = _make_item(
                row.get("word"),
                row.get("english"),
                row.get("chinese"),
                row.get("sense"),
                row.get("index", 1),
            )
            if it:
                items.append(it)
    return items


def import_sentences_excel(content: bytes) -> list[SentenceItem]:
    """解析导出的 Excel：表头 word / english / chinese / sense / index。"""
    from openpyxl import load_workbook

    try:
        wb = load_workbook(BytesIO(content), read_only=True, data_only=True)
    except Exception as e:
        raise ValueError(f"Excel 读取失败：{e}") from e

    items: list[SentenceItem] = []
    try:
        for ws in wb.worksheets:
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                continue
            header = [str(c).strip().lower() if c is not None else "" for c in rows[0]]
            col = {name: i for i, name in enumerate(header)}
            has_header = "word" in col and "english" in col

            def cell(row: tuple, name: str) -> object:
                i = col.get(name, -1)
                return row[i] if 0 <= i < len(row) else ""

            data_rows = rows[1:] if has_header else rows
            for row in data_rows:
                if has_header:
                    it = _make_item(
                        cell(row, "word"),
                        cell(row, "english"),
                        cell(row, "chinese"),
                        cell(row, "sense"),
                        cell(row, "index") or 1,
                    )
                else:
                    # 无表头时按导出顺序假设：word, english, chinese, sense, index
                    it = _make_item(
                        row[0] if len(row) > 0 else "",
                        row[1] if len(row) > 1 else "",
                        row[2] if len(row) > 2 else "",
                        row[3] if len(row) > 3 else "",
                        row[4] if len(row) > 4 and row[4] else 1,
                    )
                if it:
                    items.append(it)
    finally:
        wb.close()
    return items


_MD_WORD_RE = re.compile(r"^#{2,6}\s+(.+?)\s*$")  # ## word（## 及以上为词标题）
_MD_TITLE_RE = re.compile(r"^#\s+(.+?)\s*$")  # # 顶层标题，跳过
_MD_SENT_RE = re.compile(r"^\s*(\d+)[\.、\)]\s+\*\*(.+?)\*\*\s*$")  # 1. **english**
_MD_TRANS_RE = re.compile(r"^\s*[-*]\s*译文[：:]\s*(.*)$")
_MD_SENSE_RE = re.compile(r"^\s*[-*]\s*词义[：:]\s*(.*)$")


def import_sentences_markdown(content: bytes) -> list[SentenceItem]:
    """解析导出的 Markdown（词为 ## 标题，句子为 `N. **english**`）。"""
    try:
        text = content.decode("utf-8-sig")
    except UnicodeDecodeError as e:
        raise ValueError(f"文件编码无法识别：{e}") from e

    items: list[SentenceItem] = []
    word = ""
    english = chinese = sense = ""
    index = 1

    def flush() -> None:
        nonlocal english, chinese, sense, index
        it = _make_item(word, english, chinese, sense, index)
        if it:
            items.append(it)
        english = chinese = sense = ""
        index = 1

    for line in text.splitlines():
        if _MD_TITLE_RE.match(line):  # 顶层标题（# 生词例句复习）
            continue
        m = _MD_WORD_RE.match(line)
        if m:
            flush()  # 提交上一个词的最后一句
            word = m.group(1).strip()
            continue
        m = _MD_SENT_RE.match(line)
        if m:
            flush()  # 上一句结束
            index = int(m.group(1))
            english = m.group(2).strip()
            continue
        m = _MD_TRANS_RE.match(line)
        if m:
            chinese = m.group(1).strip()
            continue
        m = _MD_SENSE_RE.match(line)
        if m:
            sense = m.group(1).strip()
            continue
    flush()

    return items


def import_sentences_file(filename: str, content: bytes) -> list[SentenceItem]:
    """根据扩展名自动选择例句导入器。"""
    suffix = Path(filename or "").suffix.lower()
    if suffix == ".json":
        return import_sentences_json(content)
    if suffix in {".xlsx", ".xlsm"}:
        return import_sentences_excel(content)
    if suffix in {".md", ".markdown"}:
        return import_sentences_markdown(content)
    raise ValueError("暂不支持该格式。请使用本应用导出的 Excel（.xlsx）/ JSON / Markdown 文件。")
